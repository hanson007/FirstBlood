#! /usr/bin/env python
# -*-coding:utf-8-*-
##################################################
# Function:        银谷在线注册统计及出借统计脚本
# Usage:                python start.py
# Author:               黄小雪
# Date:				   2016年7月19日
# Company:
# Version:        1.2
##################################################

import xlwt
import xlrd
from unicode_width import *
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment


def set_style(name, height, bold=False):
    # 设置单元格样式
    style = xlwt.XFStyle()  # 初始化样式

    font = xlwt.Font()  # 为样式创建字体
    font.name = name  # 'Times New Roman'
    font.bold = bold
    font.color_index = 4
    font.height = height

    al = xlwt.Alignment()
    al.horz = xlwt.Alignment.HORZ_CENTER  # 设置水平居中
    al.vert = xlwt.Alignment.VERT_CENTER  # 设置垂直居中

    style.font = font
    style.alignment = al
    return style


def get_table(file, table=False):
    # 获取表格数据
    data = xlrd.open_workbook(file)
    sheets = False
    if table:
        table = data.sheets()[0]
    else:
        sheets = data.sheets()

    return table or sheets


def sheet_write(f, sheet_name, row0, rows, width):
    # 写入工作簿
    sheet = f.add_sheet(sheet_name, cell_overwrite_ok=True)  # 创建sheet

    # 生成第一行
    for i in range(0, len(row0)):
        sheet.write(0, i, row0[i], set_style('Times New Roman', 220, True))
        sheet.col(i).width = 256 * width[i]

    for j in range(0, len(rows)):
        row = rows[j]
        for i in range(0, len(row)):
            sheet.write(j + 1, i, row[i], set_style('Times New Roman', 220, False))


class Openpyxl(object):
    # openpyxl 生成excel文件
    def __init__(self, filename):
        self.ft1 = Font(name='Calibri', bold=True)
        self.ft2 = Font(name='Calibri')
        self.al = Alignment(horizontal='center', vertical='center')
        self.filename = filename
        self.wb = Workbook()

    def __set_width(self, rows, ws):
        widths = get_width(*rows)
        for i in xrange(len(widths)):
            ws.column_dimensions[get_column_letter(i + 1)].width = widths[i]

    def __get_new_ws(self, title):
        new_ws = self.wb.get_active_sheet()

        if len(new_ws.get_cell_collection()) == 0:
            new_ws.title = title
        else:
            new_ws = self.wb.create_sheet(title=title)

        return new_ws

    def add_sheet(self, title, rows):
        new_ws = self.__get_new_ws(title)

        for i in range(len(rows)):
            row = rows[i]
            for j in range(len(row)):
                new_ws.cell(row=i + 1, column=j + 1).value = row[j]

                if i == 0:
                    new_ws.cell(row=i + 1, column=j + 1).font = self.ft1
                else:
                    new_ws.cell(row=i + 1, column=j + 1).font = self.ft2
                new_ws.cell(row=i + 1, column=j + 1).alignment = self.al

        self.__set_width(rows, new_ws)

    def save(self):
        self.wb.save(filename=self.filename)
